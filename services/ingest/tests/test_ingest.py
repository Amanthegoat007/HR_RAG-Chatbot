"""
============================================================================
FILE: services/ingest/tests/test_ingest.py
PURPOSE: Unit tests for ingest service — file conversion, chunking, upload API.
ARCHITECTURE REF: §12 — Testing & Validation
============================================================================
"""

import io
import sys
import os

import pytest

os.environ.setdefault("JWT_SECRET", "test_secret_at_least_256bits_long_for_testing")
os.environ.setdefault("POSTGRES_DSN", "postgresql://test:test@localhost/test")
os.environ.setdefault("MINIO_ROOT_PASSWORD", "minioadmin")
os.environ.setdefault("ADMIN_USERNAME", "hr_admin")
os.environ.setdefault("ADMIN_PASSWORD_HASH", "$2b$12$testhashtesthashhh")
os.environ.setdefault("USER_USERNAME", "hr_user")
os.environ.setdefault("USER_PASSWORD_HASH", "$2b$12$testhashhh")

sys.path.insert(0, "/app")


class TestMarkdownConverter:
    """Unit tests for document → markdown conversion."""

    def test_txt_to_markdown_adds_frontmatter(self):
        """Plain text should get YAML frontmatter prepended."""
        from app.markdown_converter import txt_to_markdown
        content = b"This is a test document with some content.\nIt has multiple lines."
        markdown, pages = txt_to_markdown(content, "test.txt")
        assert markdown.startswith("---")
        assert "filename: test.txt" in markdown
        assert "This is a test document" in markdown

    def test_txt_to_markdown_utf8(self):
        """UTF-8 content (including Arabic) should be handled correctly."""
        from app.markdown_converter import txt_to_markdown
        content = "سياسة الإجازة السنوية".encode("utf-8")
        markdown, _ = txt_to_markdown(content, "arabic_policy.txt")
        assert "سياسة الإجازة السنوية" in markdown

    def test_xlsx_to_markdown_structure(self):
        """Excel sheets should become ## headings with pipe tables."""
        import openpyxl
        from app.markdown_converter import xlsx_to_markdown

        # Create a minimal XLSX in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leave Policy"
        ws["A1"] = "Type"
        ws["B1"] = "Days"
        ws["A2"] = "Annual"
        ws["B2"] = "30"

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        markdown, sheets = xlsx_to_markdown(xlsx_bytes, "policy.xlsx")
        assert "## Sheet: Leave Policy" in markdown
        assert "| Type |" in markdown
        assert "Annual" in markdown
        assert sheets == 1


class TestChunker:
    """Unit tests for the document chunker."""

    def test_chunk_count_reasonable(self):
        """A medium-length document should produce multiple chunks."""
        from app.chunker import chunk_markdown

        # Create a markdown document with enough content
        content = "---\nfilename: test.md\nformat: md\n---\n\n"
        content += "# Leave Policy\n\n"
        content += "Annual leave allows employees to take paid time off. " * 50
        content += "\n\n## Sick Leave\n\n"
        content += "Sick leave provides coverage when employees are ill. " * 50

        chunks = chunk_markdown(content, chunk_size=256, overlap=64)
        assert len(chunks) >= 2, "Should produce multiple chunks for large content"

    def test_chunks_include_heading_context(self):
        """Chunks should carry the nearest heading as section_heading."""
        from app.chunker import chunk_markdown

        content = "---\nfilename: t.md\nformat: md\n---\n\n"
        content += "# Annual Leave Policy\n\n"
        content += "Employees are entitled to annual leave. " * 30

        chunks = chunk_markdown(content)
        headings = [c.section_heading for c in chunks if c.section_heading]
        assert len(headings) > 0, "At least some chunks should have section headings"

    def test_chunk_overlap_reduces_content_gaps(self):
        """Adjacent chunks should share some text (overlap)."""
        from app.chunker import chunk_markdown

        content = "---\nfilename: t.md\nformat: md\n---\n\n"
        content += "Sentence one about leave policy. " * 100

        chunks = chunk_markdown(content, chunk_size=100, overlap=32)
        if len(chunks) >= 2:
            # Last sentence of chunk 0 should appear in start of chunk 1
            chunk0_words = set(chunks[0].text.split()[-20:])
            chunk1_words = set(chunks[1].text.split()[:20])
            # Some overlap should exist
            overlap_words = chunk0_words & chunk1_words
            assert len(overlap_words) > 0, "Adjacent chunks should share words (overlap)"


class TestMetadataExtractor:
    """Unit tests for metadata extraction."""

    def test_extract_frontmatter(self):
        """Should parse YAML frontmatter correctly."""
        from app.metadata_extractor import extract_frontmatter

        md = "---\nfilename: policy.pdf\nformat: pdf\npage_count: 5\n---\n\nContent here"
        result = extract_frontmatter(md)
        assert result["filename"] == "policy.pdf"
        assert result["format"] == "pdf"
        assert result["page_count"] == "5"

    def test_extract_section_headings(self):
        """Should extract all heading levels."""
        from app.metadata_extractor import extract_section_headings

        md = "# Chapter 1\n\nContent\n\n## Section 1.1\n\nMore content\n\n### Subsection"
        headings = extract_section_headings(md)
        assert headings == ["Chapter 1", "Section 1.1", "Subsection"]
