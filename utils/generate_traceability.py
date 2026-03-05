"""
============================================================================
FILE: utils/generate_traceability.py
PURPOSE: Generate the HR RAG Chatbot traceability matrix as an Excel file.
         Maps architecture document sections → services → source files → tests.
ARCHITECTURE REF: §1 — Architecture Overview; §12 — Testing & Validation
DEPENDENCIES: openpyxl (install: pip install openpyxl)
USAGE: python utils/generate_traceability.py
       Output: HR_RAG_Traceability_Matrix.xlsx (in project root)
============================================================================
"""

import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


# ─── Style Constants ───────────────────────────────────────────────────────────

# Header row colors
HEADER_BG      = "1F3864"   # Dark navy blue
HEADER_FG      = "FFFFFF"   # White
SECTION_BG     = "2E75B6"   # Medium blue
SECTION_FG     = "FFFFFF"
SUBHEADER_BG   = "BDD7EE"   # Light blue
SUBHEADER_FG   = "000000"
ALT_ROW_BG     = "EBF3FB"   # Very light blue
STATUS_DONE    = "70AD47"   # Green
STATUS_PARTIAL = "FFC000"   # Amber
STATUS_PENDING = "FF0000"   # Red

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def make_header_font(bold=True, color=HEADER_FG, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)


def make_fill(hex_color):
    return PatternFill(patternType="solid", fgColor=hex_color)


def set_column_widths(ws, widths: dict):
    """Set column widths by letter → width in characters."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_header_row(ws, row_num: int, values: list, bg_color=HEADER_BG, fg_color=HEADER_FG):
    """Write and style a header row."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = make_header_font(color=fg_color)
        cell.fill = make_fill(bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_data_row(ws, row_num: int, values: list, alternate: bool = False):
    """Write a data row with optional alternating background."""
    bg = ALT_ROW_BG if alternate else "FFFFFF"
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = make_fill(bg)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER


def color_status_cell(ws, row_num: int, col_idx: int, status: str):
    """Color-code a status cell based on its value."""
    cell = ws.cell(row=row_num, column=col_idx)
    if status == "Implemented":
        cell.fill = make_fill(STATUS_DONE)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    elif status == "Partial":
        cell.fill = make_fill(STATUS_PARTIAL)
        cell.font = Font(name="Calibri", bold=True, size=9)
    elif status == "Pending":
        cell.fill = make_fill(STATUS_PENDING)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ─── Sheet 1: Requirements Traceability ──────────────────────────────────────

REQUIREMENTS = [
    # (Req ID, Arch Section, Requirement, Service(s), Files, Test Coverage, Status)
    ("REQ-001", "§2.1", "JWT HS256 authentication with 8h token expiry",
     "auth-svc", "services/auth/app/jwt_handler.py\nservices/auth/app/auth_service.py",
     "services/auth/tests/test_auth.py::TestLoginEndpoint", "Implemented"),

    ("REQ-002", "§2.2", "bcrypt password hashing (cost=12) for stored credentials",
     "auth-svc", "services/auth/app/auth_service.py",
     "services/auth/tests/test_auth.py::test_wrong_password_returns_401", "Implemented"),

    ("REQ-003", "§2.3", "Audit log for all login attempts (success and failure)",
     "auth-svc", "services/auth/app/auth_service.py\ndb/init.sql (audit_log table)",
     "services/auth/tests/test_auth.py", "Implemented"),

    ("REQ-004", "§3.1", "Semantic cache with cosine similarity ≥ 0.92 threshold",
     "query-svc", "services/query/app/cache.py",
     "services/query/tests/test_pipeline.py::TestSemanticCacheLogic", "Implemented"),

    ("REQ-005", "§3.2", "Redis-backed semantic cache with LRU eviction (1000 entries)",
     "query-svc", "services/query/app/cache.py",
     "services/query/tests/test_pipeline.py::TestSemanticCacheLogic", "Implemented"),

    ("REQ-006", "§3.3", "BGE-M3 embedding (1024-dim dense + sparse) via FlagEmbedding",
     "embedding-svc", "services/embed/app/embedding_service.py\nservices/embed/app/main.py",
     "services/embed/tests/test_embed.py", "Implemented"),

    ("REQ-007", "§3.4", "Batch embedding with OMP_NUM_THREADS=4 for CPU optimization",
     "embedding-svc", "services/embed/app/embedding_service.py\nservices/embed/Dockerfile",
     "services/embed/tests/test_embed.py", "Implemented"),

    ("REQ-008", "§3.5", "Hybrid retrieval: dense ANN + sparse BM25 in parallel",
     "query-svc", "services/query/app/retriever.py",
     "services/query/tests/test_pipeline.py", "Implemented"),

    ("REQ-009", "§3.5", "Reciprocal Rank Fusion (RRF k=60) for merging search results",
     "query-svc", "services/query/app/retriever.py",
     "services/query/tests/test_pipeline.py", "Implemented"),

    ("REQ-010", "§3.6", "BGE-Reranker-v2-m3 cross-encoder reranking of top-20 candidates",
     "reranker-svc, query-svc", "services/rerank/app/reranker_service.py\nservices/query/app/reranker_client.py",
     "services/rerank/tests/test_rerank.py", "Implemented"),

    ("REQ-011", "§3.7", "Local LLM (Mistral-7B Q5_K_M) via llama.cpp HTTP server",
     "llm-server, query-svc", "services/llm/Dockerfile\nservices/query/app/llm_client.py",
     "services/query/tests/test_pipeline.py::TestCircuitBreaker", "Implemented"),

    ("REQ-012", "§3.7", "Azure OpenAI fallback with circuit breaker (3 failures / 60s)",
     "query-svc", "services/query/app/llm_client.py",
     "services/query/tests/test_pipeline.py::TestCircuitBreaker", "Implemented"),

    ("REQ-013", "§3.8", "SSE streaming response: token events + sources event + done event",
     "query-svc", "services/query/app/sse_handler.py\nservices/query/app/main.py",
     "services/query/tests/test_pipeline.py::TestSSEHandler", "Implemented"),

    ("REQ-014", "§4.1", "Convert-to-Markdown: PDF, DOCX, XLSX, PPTX, TXT → Markdown",
     "ingest-svc", "services/ingest/app/markdown_converter.py\nservices/ingest/app/file_converter.py",
     "services/ingest/tests/test_ingest.py::TestMarkdownConverter", "Implemented"),

    ("REQ-015", "§4.2", "Tesseract OCR fallback for scanned PDFs (< 50 chars/page)",
     "ingest-svc", "services/ingest/app/file_converter.py\nservices/ingest/Dockerfile (tesseract-ocr-ara)",
     "services/ingest/tests/test_ingest.py", "Implemented"),

    ("REQ-016", "§4.3", "256-token semantic chunks with 64-token overlap (sentence-aligned)",
     "ingest-svc", "services/ingest/app/chunker.py",
     "services/ingest/tests/test_ingest.py::TestChunker", "Implemented"),

    ("REQ-017", "§4.4", "Section heading context carried in every chunk (heading_path)",
     "ingest-svc", "services/ingest/app/chunker.py",
     "services/ingest/tests/test_ingest.py::TestChunker::test_chunks_include_heading_context", "Implemented"),

    ("REQ-018", "§4.5", "Async Celery worker with Redis Streams broker for document processing",
     "ingest-worker", "services/ingest/app/celery_app.py\nservices/ingest/app/tasks.py",
     "services/ingest/tests/test_ingest.py", "Implemented"),

    ("REQ-019", "§4.6", "Deterministic Qdrant point IDs (uuid5) for safe task retry",
     "ingest-svc", "services/ingest/app/tasks.py",
     "services/ingest/tests/test_ingest.py", "Implemented"),

    ("REQ-020", "§5.1", "MinIO S3 bucket auto-creation at service startup",
     "ingest-svc", "services/ingest/app/minio_client.py\nservices/ingest/app/main.py",
     "services/ingest/tests/test_ingest.py", "Implemented"),

    ("REQ-021", "§5.2", "Qdrant collection auto-creation (dense 1024-dim + sparse vectors)",
     "ingest-svc", "services/ingest/app/qdrant_client_wrapper.py",
     "services/ingest/tests/test_ingest.py", "Implemented"),

    ("REQ-022", "§6.1", "PostgreSQL schema: documents, ingestion_jobs, audit_log tables",
     "postgres", "db/init.sql",
     "services/auth/tests/test_auth.py (audit_log)", "Implemented"),

    ("REQ-023", "§7.1", "Nginx SSL termination with TLS 1.2+, rate limit 10 req/s/IP",
     "nginx", "nginx/nginx.conf",
     "Manual: curl -k https://localhost/health", "Implemented"),

    ("REQ-024", "§7.2", "SSE proxy headers: proxy_buffering off, proxy_read_timeout 300s",
     "nginx", "nginx/nginx.conf",
     "Manual: EventSource test", "Implemented"),

    ("REQ-025", "§8.1", "Docker Compose resource limits: llm-server 12G, embed 2G, rerank 2G",
     "docker-compose", "docker-compose.yml",
     "docker compose config --quiet", "Implemented"),

    ("REQ-026", "§8.2", "Healthcheck-based depends_on for correct service startup order",
     "docker-compose", "docker-compose.yml",
     "docker compose up (observe startup order)", "Implemented"),

    ("REQ-027", "§9.1", "Structured JSON logging with correlation_id for distributed tracing",
     "All services", "shared/logging_config.py",
     "Manual: docker compose logs | jq '.correlation_id'", "Implemented"),

    ("REQ-028", "§10.1", "Prometheus scraping from all 6 application services",
     "prometheus", "monitoring/prometheus/prometheus.yml",
     "http://localhost:9090/targets", "Implemented"),

    ("REQ-029", "§10.2", "Pre-built Grafana dashboard: latency, cache hits, queue depth, RAM",
     "grafana", "monitoring/grafana/provisioning/dashboards/hr-chatbot.json",
     "http://localhost:3000", "Implemented"),

    ("REQ-030", "§11.1", "Arabic language support (tesseract-ocr-ara, UTF-8 throughout)",
     "ingest-svc", "services/ingest/Dockerfile\nservices/ingest/app/markdown_converter.py",
     "services/ingest/tests/test_ingest.py::test_txt_to_markdown_utf8", "Implemented"),
]


def build_requirements_sheet(wb):
    ws = wb.create_sheet("Requirements Traceability")
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "HR RAG Chatbot — Requirements Traceability Matrix"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = make_fill(HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["Req ID", "Arch Section", "Requirement Description",
               "Service(s)", "File(s)", "Test Coverage", "Status"]
    apply_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, row_data in enumerate(REQUIREMENTS, start=1):
        row_num = i + 2
        write_data_row(ws, row_num, list(row_data), alternate=(i % 2 == 0))
        # Color-code status column (index 7 = column G)
        status = row_data[6]
        color_status_cell(ws, row_num, 7, status)
        ws.row_dimensions[row_num].height = 40

    set_column_widths(ws, {
        "A": 10, "B": 13, "C": 40, "D": 18, "E": 45, "F": 45, "G": 14,
    })
    return ws


# ─── Sheet 2: Service Inventory ───────────────────────────────────────────────

SERVICES = [
    ("nginx",          "nginx:1.27-alpine",          "80, 443",   "frontend",               "128M",  "curl -f http://localhost/health"),
    ("auth-svc",       "python:3.12-slim (custom)",  "8001",      "frontend, backend",       "512M",  "curl http://localhost:8001/health"),
    ("query-svc",      "python:3.12-slim (custom)",  "8002",      "frontend, backend",       "1G",    "curl http://localhost:8002/health"),
    ("ingest-svc",     "python:3.12-slim (custom)",  "8003",      "frontend, backend",       "1G",    "curl http://localhost:8003/health"),
    ("ingest-worker",  "python:3.12-slim (custom)",  "none",      "backend",                "1G",    "docker exec ingest-worker celery inspect ping"),
    ("embedding-svc",  "python:3.12-slim (custom)",  "8004",      "backend",                "2G",    "curl http://embedding-svc:8004/health"),
    ("reranker-svc",   "python:3.12-slim (custom)",  "8005",      "backend",                "2G",    "curl http://reranker-svc:8005/health"),
    ("llm-server",     "ubuntu:24.04 (custom)",      "8080",      "backend",                "12G",   "curl http://llm-server:8080/health"),
    ("postgres",       "postgres:16-alpine",         "5432",      "backend",                "1G",    "pg_isready -U hr_user"),
    ("redis",          "redis:7.2-alpine",           "6379",      "backend",                "512M",  "redis-cli ping"),
    ("minio",          "minio/minio:RELEASE.2024-01-01T00-00-00Z", "9000, 9001", "backend", "1G",    "curl http://minio:9000/minio/health/live"),
    ("qdrant",         "qdrant/qdrant:v1.9.0",       "6333, 6334","backend",                "2G",    "curl http://qdrant:6333/healthz"),
    ("prometheus",     "prom/prometheus:v2.50.1",    "9090",      "monitoring",             "512M",  "curl http://prometheus:9090/-/healthy"),
    ("grafana",        "grafana/grafana:10.3.1",     "3000",      "monitoring, frontend",   "256M",  "curl http://grafana:3000/api/health"),
]


def build_service_inventory_sheet(wb):
    ws = wb.create_sheet("Service Inventory")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Service Inventory — 14 Containers"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = make_fill(HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Service Name", "Docker Image", "Port(s)", "Network(s)", "Memory Limit", "Health Check"]
    apply_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    for i, row_data in enumerate(SERVICES, start=1):
        row_num = i + 2
        write_data_row(ws, row_num, list(row_data), alternate=(i % 2 == 0))
        ws.row_dimensions[row_num].height = 30

    set_column_widths(ws, {
        "A": 18, "B": 38, "C": 16, "D": 22, "E": 14, "F": 45,
    })
    return ws


# ─── Sheet 3: API Endpoints ────────────────────────────────────────────────────

API_ENDPOINTS = [
    ("POST",   "/auth/login",                 "auth-svc",   False, "any",   "Authenticate and receive JWT token"),
    ("GET",    "/auth/health",                "auth-svc",   False, "any",   "Auth service health check"),
    ("GET",    "/auth/metrics",               "auth-svc",   False, "any",   "Prometheus metrics for auth-svc"),
    ("POST",   "/api/query",                  "query-svc",  True,  "any",   "Submit HR query (SSE streaming response)"),
    ("GET",    "/api/health",                 "query-svc",  False, "any",   "Query service health check"),
    ("GET",    "/api/metrics",                "query-svc",  False, "any",   "Prometheus metrics for query-svc"),
    ("POST",   "/ingest/upload",              "ingest-svc", True,  "admin", "Upload HR document for processing"),
    ("GET",    "/ingest/documents",           "ingest-svc", True,  "any",   "List all ingested documents"),
    ("GET",    "/ingest/document/{id}",       "ingest-svc", True,  "any",   "Get document details and ingestion status"),
    ("DELETE", "/ingest/document/{id}",       "ingest-svc", True,  "admin", "Delete document and its vectors"),
    ("GET",    "/ingest/health",              "ingest-svc", False, "any",   "Ingest service health check"),
    ("GET",    "/ingest/metrics",             "ingest-svc", False, "any",   "Prometheus metrics for ingest-svc"),
    ("POST",   "/embed",                      "embed-svc",  False, "any",   "Generate dense+sparse embedding (internal)"),
    ("GET",    "/embed/health",               "embed-svc",  False, "any",   "Embedding service health check"),
    ("POST",   "/rerank",                     "rerank-svc", False, "any",   "Cross-encoder reranking (internal)"),
    ("GET",    "/rerank/health",              "rerank-svc", False, "any",   "Reranker service health check"),
    ("POST",   "/v1/chat/completions",        "llm-server", False, "any",   "LLM text generation (llama.cpp OpenAI compat)"),
    ("GET",    "/health",                     "llm-server", False, "any",   "LLM server health check"),
]


def build_api_endpoints_sheet(wb):
    ws = wb.create_sheet("API Endpoints")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "API Endpoints — All Services"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = make_fill(HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Method", "Endpoint", "Service", "Auth Required", "Role", "Description"]
    apply_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    for i, row_data in enumerate(API_ENDPOINTS, start=1):
        row_num = i + 2
        method, endpoint, service, auth_required, role, description = row_data
        write_data_row(ws, row_num,
                       [method, endpoint, service, "Yes" if auth_required else "No", role, description],
                       alternate=(i % 2 == 0))
        ws.row_dimensions[row_num].height = 22

        # Color HTTP method cells
        method_cell = ws.cell(row=row_num, column=1)
        method_colors = {"POST": "0070C0", "GET": "70AD47", "DELETE": "FF0000", "PUT": "FFC000"}
        color = method_colors.get(method, "000000")
        method_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        method_cell.fill = make_fill(color)
        method_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_column_widths(ws, {
        "A": 10, "B": 35, "C": 16, "D": 14, "E": 10, "F": 50,
    })
    return ws


# ─── Sheet 4: Data Flow ────────────────────────────────────────────────────────

DATA_FLOWS = [
    ("1",  "User Browser",    "nginx",           "HTTPS",   "QueryRequest JSON",
     "SSL terminated, rate limited, forwarded to query-svc"),
    ("2",  "nginx",           "query-svc",       "HTTP",    "QueryRequest JSON",
     "JWT validated, correlation_id set"),
    ("3",  "query-svc",       "redis",           "Redis",   "query_embedding (bytes)",
     "Semantic cache lookup using cosine similarity"),
    ("4",  "query-svc",       "embedding-svc",   "HTTP",    "EmbedRequest {text}",
     "Get dense (1024-dim) + sparse vectors for query"),
    ("5",  "query-svc",       "qdrant",          "HTTP",    "HybridQuery {dense, sparse}",
     "Prefetch + RRF fusion → top-20 chunk candidates"),
    ("6",  "query-svc",       "reranker-svc",    "HTTP",    "RerankRequest {query, docs}",
     "Cross-encoder scoring → top-5 reranked chunks"),
    ("7",  "query-svc",       "llm-server",      "HTTP SSE","ChatRequest {messages, stream:true}",
     "Streaming token generation from Mistral-7B"),
    ("8",  "query-svc",       "User Browser",    "SSE",     "token/sources/done events",
     "Forward tokens as SSE events; send sources at end"),
    ("9",  "query-svc",       "redis",           "Redis",   "answer + embeddings",
     "Cache answer for future identical/similar queries"),
    ("10", "User Browser",    "nginx",           "HTTPS",   "UploadRequest {file}",
     "HR admin uploads document (max 200MB)"),
    ("11", "nginx",           "ingest-svc",      "HTTP",    "Multipart form data",
     "JWT validated (admin role required)"),
    ("12", "ingest-svc",      "minio",           "S3 API",  "Original file bytes",
     "Store original document in hr-documents bucket"),
    ("13", "ingest-svc",      "postgres",        "asyncpg", "INSERT documents row",
     "Create document record with status=pending"),
    ("14", "ingest-svc",      "redis",           "Celery",  "process_document task",
     "Dispatch Celery task to ingest-worker"),
    ("15", "ingest-worker",   "minio",           "S3 API",  "Download original file",
     "Worker downloads file for processing"),
    ("16", "ingest-worker",   "minio",           "S3 API",  "Upload markdown bytes",
     "Store converted markdown for debugging/audit"),
    ("17", "ingest-worker",   "embedding-svc",   "HTTP",    "EmbedRequest {batch texts}",
     "Embed all chunks in batches of 32"),
    ("18", "ingest-worker",   "qdrant",          "HTTP",    "UpsertPoints {vectors, payloads}",
     "Store dense+sparse vectors with chunk metadata"),
    ("19", "ingest-worker",   "postgres",        "asyncpg", "UPDATE documents status=ready",
     "Mark document as available for querying"),
    ("20", "prometheus",      "All services",    "HTTP",    "GET /metrics",
     "Scrape Prometheus exposition format metrics"),
]


def build_data_flow_sheet(wb):
    ws = wb.create_sheet("Data Flow")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Data Flow — Service Interactions"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = make_fill(HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Step", "Source Service", "Target Service", "Protocol", "Data Payload", "Description"]
    apply_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    for i, row_data in enumerate(DATA_FLOWS, start=1):
        row_num = i + 2
        write_data_row(ws, row_num, list(row_data), alternate=(i % 2 == 0))
        ws.row_dimensions[row_num].height = 30

    set_column_widths(ws, {
        "A": 6, "B": 18, "C": 18, "D": 12, "E": 30, "F": 45,
    })
    return ws


# ─── Sheet 5: File Index ───────────────────────────────────────────────────────

FILE_INDEX = [
    # (File path, Layer, Purpose, Lines ~, Key functions/classes)
    (".env.example",                              "L0", "All environment variable definitions with documentation", 120, "—"),
    ("docker-compose.yml",                         "L0", "All 14 containers, networks, volumes, resource limits", 220, "—"),
    ("docker-compose.dev.yml",                     "L0", "Dev overrides: debug ports, hot-reload mounts", 50, "—"),
    ("docker-compose.prod.yml",                    "L0", "Prod hardening: restart always, read-only FS", 60, "—"),
    ("Makefile",                                   "L0", "Linux/WSL make targets", 80, "up, down, test, logs"),
    ("make.ps1",                                   "L0", "Windows PowerShell equivalents", 120, "up, down, test, logs"),
    ("db/init.sql",                                "L1", "PostgreSQL schema: 3 tables, 7 indexes, triggers", 120, "documents, ingestion_jobs, audit_log"),
    ("shared/logging_config.py",                   "L1", "Structured JSON/text logging with correlation_id", 80, "setup_logging(), set_correlation_id()"),
    ("shared/jwt_utils.py",                        "L1", "Shared JWT validation for all services", 70, "create_access_token(), decode_token()"),
    ("nginx/nginx.conf",                           "L1", "SSL, rate limiting, SSE proxy, security headers", 130, "—"),
    ("services/embed/app/embedding_service.py",    "L2", "BGE-M3 load + dense/sparse inference", 120, "EmbeddingService, embed_texts()"),
    ("services/embed/app/main.py",                 "L2", "Embedding FastAPI app + /embed + /health", 80, "POST /embed"),
    ("services/rerank/app/reranker_service.py",    "L2", "BGE-Reranker-v2-m3 cross-encoder scoring", 90, "RerankService, rerank()"),
    ("services/auth/app/auth_service.py",          "L3", "bcrypt verification + audit log writes", 80, "authenticate_user(), write_audit_log()"),
    ("services/auth/app/jwt_handler.py",           "L3", "JWT creation and validation", 60, "create_access_token(), decode_token()"),
    ("services/auth/app/main.py",                  "L3", "Auth FastAPI app + POST /auth/login", 80, "POST /auth/login"),
    ("services/ingest/app/markdown_converter.py",  "L4", "PDF/DOCX/XLSX/PPTX → Markdown with frontmatter", 280, "pdf_to_markdown(), docx_to_markdown()"),
    ("services/ingest/app/chunker.py",             "L4", "256-token semantic chunks, 64-token overlap", 180, "chunk_markdown(), DocumentChunk"),
    ("services/ingest/app/tasks.py",               "L4", "Celery tasks: process_document, delete_document", 200, "process_document(), delete_document()"),
    ("services/ingest/app/main.py",                "L4", "Ingest FastAPI app + upload/list/delete endpoints", 150, "POST /ingest/upload"),
    ("services/query/app/cache.py",                "L5", "Redis semantic cache with cosine similarity", 150, "SemanticCache, get(), set()"),
    ("services/query/app/retriever.py",            "L5", "Hybrid Qdrant search with RRF fusion", 80, "hybrid_search()"),
    ("services/query/app/reranker_client.py",      "L5", "HTTP client for reranker-svc", 80, "rerank_chunks()"),
    ("services/query/app/llm_client.py",           "L5", "LLM client: local + Azure fallback + circuit breaker", 160, "generate_stream(), CircuitBreaker"),
    ("services/query/app/sse_handler.py",          "L5", "SSE event builders for token/sources/error/done", 100, "make_token_event(), build_query_stream()"),
    ("services/query/app/pipeline.py",             "L5", "Central RAG orchestrator (steps 1–7)", 160, "run_query_pipeline()"),
    ("services/query/app/main.py",                 "L5", "Query FastAPI app + POST /query (SSE)", 150, "POST /query"),
    ("monitoring/prometheus/prometheus.yml",        "L6", "Prometheus scrape config for all services", 60, "—"),
    ("monitoring/grafana/provisioning/dashboards/hr-chatbot.json", "L6", "Pre-built Grafana dashboard JSON", 300, "—"),
    ("utils/hash_password.py",                     "L7", "bcrypt hash generator CLI", 40, "hash_password()"),
    ("utils/Health-Check.ps1",                     "L7", "Windows PowerShell health check", 80, "Test-ServiceHealth"),
    ("utils/load_test.py",                         "L7", "Concurrent load test with P50/P95 reporting", 200, "run_load_test()"),
    ("utils/generate_traceability.py",             "L7", "This script — generates the traceability matrix", 400, "build_*_sheet()"),
]


def build_file_index_sheet(wb):
    ws = wb.create_sheet("File Index")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    ws["A1"].value = f"File Index — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = make_fill(HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["File Path", "Layer", "Purpose", "~Lines", "Key Functions / Classes"]
    apply_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    for i, row_data in enumerate(FILE_INDEX, start=1):
        row_num = i + 2
        write_data_row(ws, row_num, list(row_data), alternate=(i % 2 == 0))
        ws.row_dimensions[row_num].height = 28

    set_column_widths(ws, {
        "A": 55, "B": 7, "C": 55, "D": 8, "E": 50,
    })
    return ws


# ─── Main ──────────────────────────────────────────────────────────────────────


def main():
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "HR_RAG_Traceability_Matrix.xlsx",
    )

    print("Generating HR RAG Chatbot Traceability Matrix...")
    print(f"Output: {output_path}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    print("  Building Sheet 1: Requirements Traceability...")
    build_requirements_sheet(wb)

    print("  Building Sheet 2: Service Inventory...")
    build_service_inventory_sheet(wb)

    print("  Building Sheet 3: API Endpoints...")
    build_api_endpoints_sheet(wb)

    print("  Building Sheet 4: Data Flow...")
    build_data_flow_sheet(wb)

    print("  Building Sheet 5: File Index...")
    build_file_index_sheet(wb)

    wb.save(output_path)
    print(f"\nDone! Traceability matrix saved to:")
    print(f"  {output_path}")
    print()
    print("Sheets generated:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")


if __name__ == "__main__":
    main()
